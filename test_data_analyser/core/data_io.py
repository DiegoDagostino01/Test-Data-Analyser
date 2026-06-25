from __future__ import annotations

import posixpath
from pathlib import Path
from typing import Any, Iterable, Literal, Optional, cast
from zipfile import ZipFile
import xml.etree.ElementTree as ET
import numpy as np
import pandas as pd

from .config import NUMERIC_EXTRACT_RE


XLRD_MISSING_MESSAGE = "xlrd is required to open .xls files. Install it with: pip install xlrd==1.2.0."
ExcelEngine = Literal["openpyxl", "xlrd"]
SMART_HEADER_PREVIEW_ROWS = 50
_XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_XLSX_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_XLSX_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_BUILTIN_DATE_NUMFMT_IDS = frozenset(
    {14, 15, 16, 17, 18, 19, 20, 21, 22, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 45, 46, 47, 50, 51, 52, 53, 54, 55, 56, 57, 58}
)


class _UnsupportedFastXlsx(Exception):
    """Raised when the direct XML xlsx reader should fall back to openpyxl."""


def _xlrd_module():
    try:
        import xlrd
    except ImportError as exc:
        raise ImportError(XLRD_MISSING_MESSAGE) from exc
    return xlrd


def get_excel_sheets(filepath: str | Path) -> list[str]:
    path = Path(filepath)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _xlsx_sheet_names(path)
    if suffix == ".xls":
        xlrd = _xlrd_module()
        workbook = xlrd.open_workbook(str(path), on_demand=True)
        try:
            return [str(sheet) for sheet in workbook.sheet_names()]
        finally:
            release_resources = getattr(workbook, "release_resources", None)
            if callable(release_resources):
                release_resources()
    if suffix not in {".xlsx", ".xls"}:
        return []
    return []

def _xlsx_sheet_names(path: Path) -> list[str]:
    """Return sheet names from xlsx workbook metadata without loading cells."""
    try:
        with ZipFile(path) as archive:
            workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        return [str(sheet.attrib.get("name", "")) for sheet in workbook.findall(f"{{{_XLSX_MAIN_NS}}}sheets/{{{_XLSX_MAIN_NS}}}sheet")]
    except Exception:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True)
        try:
            return [str(sheet) for sheet in workbook.sheetnames]
        finally:
            workbook.close()

def _is_blank_cell(value: object) -> bool:
    """Return True when an Excel cell should be treated as blank."""
    if pd.isna(cast(Any, value)):
        return True
    return str(value).strip() == ""

def _looks_numeric_cell(value: object) -> bool:
    """Return True when a cell can be interpreted as a numeric data value."""
    if _is_blank_cell(value):
        return False
    try:
        float(str(value).replace(",", "").strip())
        return True
    except Exception:
        return False

def _make_unique_column_names(columns: Iterable[object]) -> list[str]:
    """Make dataframe column names unique while preserving readable names."""
    seen: dict[str, int] = {}
    unique: list[str] = []
    for col in columns:
        base = str(col).strip() or "Column"
        count = seen.get(base, 0)
        unique.append(base if count == 0 else f"{base} ({count + 1})")
        seen[base] = count + 1
    return unique

def _read_excel_raw(
    path: Path,
    sheet_name: Optional[str],
    engine: ExcelEngine,
    *,
    nrows: int | None = None,
    skiprows: int | None = None,
) -> pd.DataFrame:
    """Read an Excel sheet without headers, preserving original row positions."""
    read_kwargs: dict[str, Any] = {
        "sheet_name": sheet_name or 0,
        "header": None,
        "engine": engine,
    }
    if nrows is not None:
        read_kwargs["nrows"] = nrows
    if skiprows:
        read_kwargs["skiprows"] = skiprows
    return pd.read_excel(path, **read_kwargs)

def _normalise_raw_excel_frame(raw: pd.DataFrame) -> pd.DataFrame:
    return raw.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)

def _find_excel_data_start(raw: pd.DataFrame) -> int | None:
    for idx in range(len(raw)):
        numeric_count = sum(_looks_numeric_cell(v) for v in raw.iloc[idx].tolist())
        if numeric_count >= 2:
            return idx
    return None

def _find_excel_data_start_position(raw: pd.DataFrame) -> int | None:
    """Return the original zero-based row position where data appears."""
    non_blank_rows = raw.dropna(how="all")
    for original_idx, row in non_blank_rows.iterrows():
        numeric_count = sum(_looks_numeric_cell(v) for v in row.tolist())
        if numeric_count >= 2:
            return int(original_idx)
    return None

def _build_grouped_excel_columns(header_block: pd.DataFrame, column_count: int) -> list[str]:
    if header_block.empty:
        return _make_unique_column_names([f"Column {i + 1}" for i in range(column_count)])
    header_block = header_block.ffill(axis=1)
    built_columns: list[str] = []
    for col_idx in range(column_count):
        parts: list[str] = []
        if col_idx < header_block.shape[1]:
            for row_idx in range(header_block.shape[0]):
                value = header_block.iat[row_idx, col_idx]
                if _is_blank_cell(value):
                    continue
                part = str(value).strip()
                if part and part not in parts:
                    parts.append(part)
        built_columns.append(" - ".join(parts) if parts else f"Column {col_idx + 1}")
    return _make_unique_column_names(built_columns)

def _dataframe_from_smart_header_rows(header_rows: list[list[object]], data_rows: list[list[object]]) -> pd.DataFrame:
    column_count = max((len(row) for row in [*header_rows, *data_rows]), default=0)
    if column_count:
        header_rows = [row + [None] * (column_count - len(row)) for row in header_rows]
        data_rows = [row + [None] * (column_count - len(row)) for row in data_rows]
    data = pd.DataFrame(data_rows)
    data.columns = _build_grouped_excel_columns(pd.DataFrame(header_rows), data.shape[1])
    return data.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)

def _xlsx_relationship_targets(archive: ZipFile) -> dict[str, str]:
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    return {
        str(rel.attrib["Id"]): str(rel.attrib["Target"])
        for rel in rels.findall(f"{{{_XLSX_PACKAGE_REL_NS}}}Relationship")
    }

def _xlsx_sheet_xml_path(archive: ZipFile, sheet_name: Optional[str]) -> str:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    relationships = _xlsx_relationship_targets(archive)
    sheets = workbook.findall(f"{{{_XLSX_MAIN_NS}}}sheets/{{{_XLSX_MAIN_NS}}}sheet")
    if not sheets:
        raise _UnsupportedFastXlsx("Workbook has no sheets.")
    selected = sheet_name or str(sheets[0].attrib.get("name", ""))
    for sheet in sheets:
        if str(sheet.attrib.get("name", "")) != selected:
            continue
        relationship_id = sheet.attrib.get(f"{{{_XLSX_REL_NS}}}id")
        target = relationships.get(str(relationship_id))
        if not target:
            raise _UnsupportedFastXlsx("Worksheet relationship target is missing.")
        target = target.lstrip("/")
        if target.startswith("xl/"):
            return posixpath.normpath(target)
        return posixpath.normpath(posixpath.join("xl", target))
    raise KeyError(f"Worksheet not found: {selected}")

def _xlsx_shared_strings(archive: ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    return ["".join(item.itertext()) for item in root.findall(f"{{{_XLSX_MAIN_NS}}}si")]

def _format_code_looks_date(format_code: str) -> bool:
    cleaned = ""
    in_quote = False
    escaped = False
    for char in format_code.lower():
        if escaped:
            escaped = False
            continue
        if char == "\\":
            escaped = True
            continue
        if char == '"':
            in_quote = not in_quote
            continue
        if not in_quote:
            cleaned += char
    return any(token in cleaned for token in ("yy", "dd", "hh", "ss")) or ("m" in cleaned and "0" not in cleaned.replace("m", ""))

def _xlsx_date_style_indexes(archive: ZipFile) -> set[int]:
    try:
        root = ET.fromstring(archive.read("xl/styles.xml"))
    except KeyError:
        return set()
    custom_date_ids = {
        int(num_format.attrib.get("numFmtId", -1))
        for num_format in root.findall(f"{{{_XLSX_MAIN_NS}}}numFmts/{{{_XLSX_MAIN_NS}}}numFmt")
        if _format_code_looks_date(str(num_format.attrib.get("formatCode", "")))
    }
    date_num_format_ids = set(_BUILTIN_DATE_NUMFMT_IDS) | custom_date_ids
    date_styles: set[int] = set()
    for style_index, style in enumerate(root.findall(f"{{{_XLSX_MAIN_NS}}}cellXfs/{{{_XLSX_MAIN_NS}}}xf")):
        try:
            num_format_id = int(style.attrib.get("numFmtId", -1))
        except ValueError:
            continue
        if num_format_id in date_num_format_ids:
            date_styles.add(style_index)
    return date_styles

def _xlsx_column_index(cell_ref: str) -> int | None:
    index = 0
    found = False
    for char in cell_ref:
        if not char.isalpha():
            break
        found = True
        index = index * 26 + (ord(char.upper()) - 64)
    return index - 1 if found else None

def _xlsx_cell_text(cell: ET.Element, tag: str) -> str | None:
    value = cell.find(f"{{{_XLSX_MAIN_NS}}}{tag}")
    return None if value is None else value.text

def _xlsx_cell_value(cell: ET.Element, shared_strings: list[str], date_styles: set[int]) -> object:
    cell_type = cell.attrib.get("t")
    style = cell.attrib.get("s")
    if style is not None:
        try:
            if int(style) in date_styles:
                raise _UnsupportedFastXlsx("Date-formatted cell requires openpyxl conversion.")
        except ValueError:
            pass
    if cell_type == "inlineStr":
        inline = cell.find(f"{{{_XLSX_MAIN_NS}}}is")
        return "" if inline is None else "".join(inline.itertext())
    text = _xlsx_cell_text(cell, "v")
    if text is None:
        return None
    if cell_type == "s":
        try:
            return shared_strings[int(text)]
        except (IndexError, ValueError):
            return text
    if cell_type == "b":
        return text == "1"
    if cell_type == "e":
        return None
    if cell_type in {"str", "d"}:
        if cell_type == "d":
            raise _UnsupportedFastXlsx("ISO date cell requires openpyxl conversion.")
        return text
    try:
        number = float(text)
    except ValueError:
        return text
    return int(number) if number.is_integer() else number

def _read_xlsx_with_smart_headers_xml(path: Path, sheet_name: Optional[str]) -> pd.DataFrame:
    with ZipFile(path) as archive:
        sheet_path = _xlsx_sheet_xml_path(archive, sheet_name)
        shared_strings = _xlsx_shared_strings(archive)
        date_styles = _xlsx_date_style_indexes(archive)
        header_rows: list[list[object]] = []
        data_rows: list[list[object]] = []
        data_started = False
        with archive.open(sheet_path) as sheet_xml:
            for _event, row in ET.iterparse(sheet_xml, events=("end",)):
                if row.tag != f"{{{_XLSX_MAIN_NS}}}row":
                    continue
                row_values: list[object] = []
                for cell in row.findall(f"{{{_XLSX_MAIN_NS}}}c"):
                    index = _xlsx_column_index(str(cell.attrib.get("r", "")))
                    if index is None:
                        index = len(row_values)
                    if index >= len(row_values):
                        row_values.extend([None] * (index + 1 - len(row_values)))
                    row_values[index] = _xlsx_cell_value(cell, shared_strings, date_styles)
                if not data_started:
                    numeric_count = sum(_looks_numeric_cell(value) for value in row_values)
                    if numeric_count >= 2:
                        data_started = True
                        data_rows.append(row_values)
                    elif any(not _is_blank_cell(value) for value in row_values):
                        header_rows.append(row_values)
                else:
                    data_rows.append(row_values)
                row.clear()
    if not data_started:
        raise _UnsupportedFastXlsx("Could not identify a numeric data row.")
    return _dataframe_from_smart_header_rows(header_rows, data_rows)

def _read_excel_default_headers(path: Path, sheet_name: Optional[str], engine: ExcelEngine) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name or 0, engine=engine)
    df = _normalise_raw_excel_frame(df)
    df.columns = _make_unique_column_names(df.columns)
    return df

def _read_xlsx_with_smart_headers_openpyxl(path: Path, sheet_name: Optional[str]) -> pd.DataFrame:
    """Read an xlsx sheet in one streaming pass using the smart-header rules."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        selected_sheet = sheet_name or workbook.sheetnames[0]
        sheet = workbook[selected_sheet]
        header_rows: list[list[object]] = []
        data_rows: list[list[object]] = []
        data_started = False
        for row in sheet.iter_rows(values_only=True):
            row_values = list(row)
            if not data_started:
                numeric_count = sum(_looks_numeric_cell(value) for value in row_values)
                if numeric_count >= 2:
                    data_started = True
                    data_rows.append(row_values)
                elif any(not _is_blank_cell(value) for value in row_values):
                    header_rows.append(row_values)
                continue
            data_rows.append(row_values)
    finally:
        workbook.close()

    if not data_started:
        return _read_excel_default_headers(path, sheet_name, "openpyxl")

    return _dataframe_from_smart_header_rows(header_rows, data_rows)

def _settings_value(settings_manager: Any, section: str, key: str, default: Any) -> Any:
    if settings_manager is None:
        return default
    try:
        return settings_manager.get(section, key)
    except Exception:
        return default

def _read_excel_with_smart_headers(
    path: Path,
    sheet_name: Optional[str],
    engine: ExcelEngine,
    settings_manager: Any = None,
) -> pd.DataFrame:
    """
    Read Excel files with normal single-row headers or grouped/multi-row headers.
    This prevents Pandas-generated "Unnamed" columns from appearing in the GUI.
    """
    header_row = int(_settings_value(settings_manager, "data_import", "header_row_index", 0) or 0)
    skip_rows = int(_settings_value(settings_manager, "data_import", "skip_rows", 0) or 0)
    if header_row != 0 or skip_rows != 0:
        df = pd.read_excel(
            path,
            sheet_name=sheet_name or 0,
            header=header_row,
            skiprows=skip_rows,
            engine=engine,
        )
        df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
        df.columns = _make_unique_column_names(df.columns)
        return df

    if engine == "openpyxl":
        try:
            return _read_xlsx_with_smart_headers_xml(path, sheet_name)
        except (_UnsupportedFastXlsx, KeyError, OSError, ValueError, ET.ParseError):
            return _read_xlsx_with_smart_headers_openpyxl(path, sheet_name)

    preview = _read_excel_raw(path, sheet_name, engine, nrows=SMART_HEADER_PREVIEW_ROWS)
    data_start_position = _find_excel_data_start_position(preview)
    if data_start_position is None:
        raw = _normalise_raw_excel_frame(_read_excel_raw(path, sheet_name, engine))
        data_start = _find_excel_data_start(raw)
        if data_start is None:
            return _read_excel_default_headers(path, sheet_name, engine)
        header_block = raw.iloc[:data_start].copy()
        data = raw.iloc[data_start:].copy().reset_index(drop=True)
        data.columns = _build_grouped_excel_columns(header_block, data.shape[1])
        data = data.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
        return data

    data = _read_excel_raw(path, sheet_name, engine, skiprows=data_start_position)
    data = data.dropna(how="all").reset_index(drop=True)
    if data.empty:
        return data

    header_block = preview.iloc[:data_start_position].dropna(how="all").copy().reset_index(drop=True)
    data.columns = _build_grouped_excel_columns(header_block, data.shape[1])
    data = data.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
    return data

def load_data(filepath: str | Path, sheet_name: Optional[str] = None, settings_manager: Any = None) -> pd.DataFrame:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    ext = path.suffix.lower()
    if ext == ".csv":
        delimiter = _settings_value(settings_manager, "data_import", "default_delimiter", "auto")
        encoding = str(_settings_value(settings_manager, "data_import", "default_encoding", "utf-8") or "utf-8")
        header_row = int(_settings_value(settings_manager, "data_import", "header_row_index", 0) or 0)
        skip_rows = int(_settings_value(settings_manager, "data_import", "skip_rows", 0) or 0)
        decimal_separator = str(_settings_value(settings_manager, "data_import", "decimal_separator", ".") or ".")
        read_kwargs: dict[str, Any] = {
            "encoding": encoding,
            "header": header_row,
            "skiprows": skip_rows,
            "decimal": decimal_separator,
        }
        if delimiter == "auto":
            read_kwargs.update({"sep": None, "engine": "python"})
        else:
            read_kwargs["sep"] = delimiter
        df = pd.read_csv(path, **read_kwargs)
        df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
        df.columns = _make_unique_column_names([str(c).strip() for c in df.columns])
        return df
    if ext == ".xlsx":
        return _read_excel_with_smart_headers(path, sheet_name, engine="openpyxl", settings_manager=settings_manager)
    if ext == ".xls":
        _xlrd_module()
        return _read_excel_with_smart_headers(path, sheet_name, engine="xlrd", settings_manager=settings_manager)

    raise ValueError("Unsupported file format. Please use CSV, XLSX, or XLS.")

def numeric_series(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce")
    text = series.astype(str).str.strip()
    text = text.replace({"": np.nan, "-": np.nan, "--": np.nan,
                         "N/A": np.nan, "n/a": np.nan, "None": np.nan})
    text = text.str.replace(",", "", regex=False)
    text = text.str.extract(NUMERIC_EXTRACT_RE, expand=False)
    return pd.to_numeric(text, errors="coerce")


#: Blank/placeholder tokens treated as missing values when classifying a column.
_BLANK_TOKENS = frozenset({"", "-", "--", "N/A", "n/a", "None", "none", "nan", "NaN"})


def detect_column_data_type(series: pd.Series, *, threshold: float = 0.6) -> str:
    """Classify a column as ``"numeric"`` or ``"text"``.

    Detection mirrors the tolerant :func:`numeric_series` coercion used by the
    plotting pipeline: a column is numeric when at least ``threshold`` of its
    non-blank values coerce to numbers. Fully blank columns are treated as
    numeric so a freshly created manual column stays plottable until
    non-numeric text is entered.
    """
    if pd.api.types.is_numeric_dtype(series):
        return "numeric"
    text = series.astype(str).str.strip()
    non_blank = ~(series.isna() | text.isin(_BLANK_TOKENS))
    total = int(non_blank.sum())
    if total == 0:
        return "numeric"
    numeric_text = text.replace({"": np.nan, "-": np.nan, "--": np.nan,
                                 "N/A": np.nan, "n/a": np.nan, "None": np.nan})
    numeric_text = numeric_text.str.replace(",", "", regex=False)
    numeric_text = numeric_text.str.extract(NUMERIC_EXTRACT_RE, expand=False)
    coerced = pd.to_numeric(numeric_text, errors="coerce")
    numeric_count = int((coerced.notna() & non_blank).sum())
    return "numeric" if (numeric_count / total) >= threshold else "text"

